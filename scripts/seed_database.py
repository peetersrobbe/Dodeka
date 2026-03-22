"""
DODEKA SPORT – Database Seeder (v2)
Re-imports all historical statistics from the Excel file into Supabase.

Sheet layout (all seasons):
  Row 1: result per match   (v=verloren, g=gelijkspel, w=gewonnen, f=forfait)
  Row 2: location per match (t=thuis, u=uit) – may be empty for 2021-2022
  Row 3: score per match    (e.g. '4-2', 'ff 5-0', 'Uitgesteld')
          Score is ALWAYS in conventional "home team – away team" format.
          For a home game:  score_dodeka   = first number
          For an away game: score_dodeka   = second number
  Row 4: opponent name
  Row 5: NAAM header + match dates
  Row 6: match times (only in seasons ≥ 2020-2021) OR first player row (older)
  Row 7+: player rows [col A = nr, col B = naam, col C+ = attendance codes]

Attendance codes:
  h = volledig   (full game)
  s = gewisseld  (substituted / partial game)
  t = toeschouwer (spectator, didn't play)
  k = volledig + is_keeper=True (played the full game as goalkeeper)
  w = old code for absent; skip (don't insert record)

Usage:
  1. pip install openpyxl supabase python-dotenv
  2. Create .env with SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY
  3. Run BOTH SQL migrations in Supabase first:
       supabase/add_keeper_to_aanwezigheid.sql
       supabase/add_player_seasons.sql
  4. python scripts/seed_database.py
     (DELETES all existing matches and re-imports everything)
"""

import os
import re
import sys
from datetime import datetime

import openpyxl

EXCEL_PATH = 'DODEKA SPORT seizoen 2025-2026 en oude statistieken.xlsx'

CURRENT_SEASON = '2025-2026'

# All season sheet names to import
SEASON_SHEETS = [
    '2017-2018', '2018-2019', '2019-2020', '2020-2021',
    '2021-2022', '2022-2023', '2023-2024', '2024-2025',
    '2025-2026',
]

# Attendance code → (status, is_keeper)
CODE_MAP = {
    'h': ('volledig',    False),
    's': ('gewisseld',   False),
    't': ('toeschouwer', False),
    'k': ('volledig',    True),   # keeper who played full game
    # 'w' is absent – we skip it (no record inserted)
}


# ── Setup ─────────────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from supabase import create_client, Client

SUPABASE_URL = os.environ['SUPABASE_URL']
SUPABASE_KEY = os.environ['SUPABASE_SERVICE_ROLE_KEY']
sb: Client   = create_client(SUPABASE_URL, SUPABASE_KEY)


# ── Player cache ──────────────────────────────────────────────
_player_cache: dict[str, str] = {}

def get_or_create_player(naam: str) -> str:
    naam = naam.strip()
    if naam in _player_cache:
        return _player_cache[naam]
    res = sb.table('players').select('id').eq('naam', naam).maybe_single().execute()
    if res.data:
        _player_cache[naam] = res.data['id']
    else:
        res = sb.table('players').insert({'naam': naam, 'active': True}).execute()
        _player_cache[naam] = res.data[0]['id']
    return _player_cache[naam]


# ── Season cache ──────────────────────────────────────────────
_season_cache: dict[str, str] = {}

def get_or_create_season(naam: str) -> str:
    if naam in _season_cache:
        return _season_cache[naam]
    res = sb.table('seasons').select('id').eq('naam', naam).maybe_single().execute()
    if res.data:
        _season_cache[naam] = res.data['id']
    else:
        res = sb.table('seasons').insert({
            'naam': naam,
            'is_current': naam == CURRENT_SEASON,
        }).execute()
        _season_cache[naam] = res.data[0]['id']
    return _season_cache[naam]


# ── Date parsing ──────────────────────────────────────────────
def parse_date(raw, season_name: str) -> str | None:
    """Return 'YYYY-MM-DD' string or None."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.strftime('%Y-%m-%d')
    s = str(raw).strip()
    if not s or s == 'None':
        return None
    if '/' in s:
        # Older format: 'DD/MM' (2017-2018 only)
        parts = s.split('/')
        try:
            day, month = int(parts[0]), int(parts[1])
            first_year = int(season_name[:4])
            year = first_year if month >= 8 else first_year + 1
            return f'{year}-{month:02d}-{day:02d}'
        except (ValueError, IndexError):
            return None
    try:
        from datetime import datetime as dt
        return dt.fromisoformat(s[:10]).strftime('%Y-%m-%d')
    except Exception:
        try:
            # Try dd/mm/yyyy or other formats
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                try:
                    return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
                except ValueError:
                    continue
        except Exception:
            pass
    return None


# ── Score parsing ─────────────────────────────────────────────
def parse_score(score_raw, is_thuis) -> tuple[int | None, int | None]:
    """
    Parse score string and return (score_dodeka, score_tegenstander).
    Score is always in conventional home-away format:
      - home game (is_thuis=True or None): first=dodeka, second=opponent
      - away game (is_thuis=False):        first=opponent, second=dodeka
    """
    if not score_raw:
        return None, None
    s = str(score_raw).strip()
    # Strip forfait prefix
    s = re.sub(r'^ff\s*', '', s, flags=re.IGNORECASE).strip()
    m = re.match(r'^(\d+)\s*[-–]\s*(\d+)$', s)
    if not m:
        return None, None
    first, second = int(m.group(1)), int(m.group(2))
    if is_thuis is False:
        # Away: conventional score is opponent(home)-dodeka(away)
        return second, first
    else:
        # Home or unknown: conventional score is dodeka-opponent
        return first, second


# ── Import one season attendance sheet ────────────────────────
def import_season_sheet(ws, season_name: str):
    """
    Parse a season attendance sheet and upsert matches + aanwezigheid.
    """
    season_id = get_or_create_season(season_name)
    print(f"  Importing {season_name}…")

    # Determine player start row:
    # If cell(6,1) is an integer → players start at row 6 (older style, no time row)
    # Otherwise → row 7 (newer style with time row in row 6)
    if isinstance(ws.cell(6, 1).value, int):
        player_start_row = 6
        time_row = None
    else:
        player_start_row = 7
        time_row = 6

    # Identify valid match columns: row 1 must have a result code (v/g/w/f)
    match_cols = []
    for c in range(3, ws.max_column + 1):
        result_val = ws.cell(1, c).value
        if result_val is None:
            continue
        result = str(result_val).strip().lower()
        if result not in ('v', 'g', 'w', 'f'):
            continue

        # Location
        loc_val = ws.cell(2, c).value
        if loc_val == 't':
            is_thuis = True
        elif loc_val == 'u':
            is_thuis = False
        else:
            is_thuis = None  # unknown (e.g. 2021-2022)

        # Opponent
        opp_val = ws.cell(4, c).value
        if not opp_val:
            continue
        opponent = str(opp_val).strip()

        # Date
        datum_str = parse_date(ws.cell(5, c).value, season_name) or '1900-01-01'

        # Time
        tijdstip = None
        if time_row:
            t_val = ws.cell(time_row, c).value
            if t_val:
                tijdstip = str(t_val).strip()

        # Score
        score_raw = str(ws.cell(3, c).value).strip() if ws.cell(3, c).value else ''
        score_d, score_t = parse_score(score_raw, is_thuis)

        is_forfait = (result == 'f')

        # Insert match
        match_payload = {
            'season_id':          season_id,
            'datum':              datum_str,
            'tegenstander':       opponent,
            'is_thuis':           is_thuis if is_thuis is not None else True,
            'score_dodeka':       score_d,
            'score_tegenstander': score_t,
            'is_forfait':         is_forfait,
            'source':             'excel',
        }
        if tijdstip:
            match_payload['tijdstip'] = tijdstip

        res = sb.table('matches').insert(match_payload).execute()
        match_id = res.data[0]['id']
        match_cols.append({'col': c, 'match_id': match_id, 'is_forfait': is_forfait})

    # Import player attendance
    aanw_batch = []
    for row in range(player_start_row, ws.max_row + 1):
        player_num = ws.cell(row, 1).value
        player_name_raw = ws.cell(row, 2).value

        # Must have a numeric player number and a name
        if not isinstance(player_num, (int, float)) or not player_name_raw:
            continue
        player_name = str(player_name_raw).strip()
        if not player_name or player_name.upper() == 'NAAM':
            continue

        player_id = get_or_create_player(player_name)

        for meta in match_cols:
            if meta['is_forfait']:
                continue  # No attendance recorded for forfait matches

            val = ws.cell(row, meta['col']).value
            if val is None:
                continue
            code = str(val).strip().lower()
            if code not in CODE_MAP:
                continue  # Skip 'w', formulas, numbers etc.

            status, is_keeper = CODE_MAP[code]
            aanw_batch.append({
                'match_id':  meta['match_id'],
                'player_id': player_id,
                'status':    status,
                'is_keeper': is_keeper,
            })

    # Bulk insert aanwezigheid in chunks of 500
    for i in range(0, len(aanw_batch), 500):
        sb.table('aanwezigheid').insert(aanw_batch[i:i+500]).execute()

    print(f"    → {len(match_cols)} matches, {len(aanw_batch)} attendance records.")


# ── Import goals (Doelpunten) sheet ──────────────────────────
def import_goals_sheet(ws, season_name: str):
    """
    Sheet layout:
      Row 1: scores   Row 2: opponents   Row 3: NAAM + dates   Row 4+: player rows
    """
    season_id = get_or_create_season(season_name)
    print(f"  Importing goals for {season_name}…")

    # Build match lookup: opponent+date → match_id
    match_lookup = {}
    for c in range(3, ws.max_column + 1):
        opp_val = ws.cell(2, c).value
        date_raw = ws.cell(3, c).value
        if not opp_val:
            continue
        opponent = str(opp_val).strip()
        datum_str = parse_date(date_raw, season_name) or '1900-01-01'
        key = (opponent, datum_str)
        if key in match_lookup:
            match_lookup[key].append(c)
        else:
            match_lookup[key] = [c]

    # Resolve match_ids
    col_to_match = {}
    for (opp, datum), cols in match_lookup.items():
        res = sb.table('matches').select('id') \
            .eq('season_id', season_id) \
            .eq('tegenstander', opp) \
            .eq('datum', datum) \
            .maybe_single().execute()
        if res.data:
            for c in cols:
                col_to_match[c] = res.data['id']

    count = 0
    for row in range(4, ws.max_row + 1):
        player_num = ws.cell(row, 1).value
        naam_raw = ws.cell(row, 2).value
        if not isinstance(player_num, (int, float)) or not naam_raw:
            continue
        naam = str(naam_raw).strip()
        if not naam:
            continue
        is_eigen = 'eigen doelpunt' in naam.lower()
        player_id = None if is_eigen else get_or_create_player(naam)

        for c, match_id in col_to_match.items():
            val = ws.cell(row, c).value
            if val is None:
                continue
            try:
                aantal = int(float(val))
            except (ValueError, TypeError):
                continue
            if aantal <= 0:
                continue
            if player_id:
                sb.table('doelpunten').insert({
                    'match_id': match_id, 'player_id': player_id, 'aantal': aantal,
                }).execute()
                count += 1

    print(f"    → {count} goal records.")


# ── Import assists sheet ──────────────────────────────────────
def import_assists_sheet(ws, season_name: str):
    season_id = get_or_create_season(season_name)
    print(f"  Importing assists for {season_name}…")

    col_to_match = {}
    for c in range(3, ws.max_column + 1):
        opp_val = ws.cell(2, c).value
        date_raw = ws.cell(3, c).value
        if not opp_val:
            continue
        opp = str(opp_val).strip()
        datum = parse_date(date_raw, season_name) or '1900-01-01'
        res = sb.table('matches').select('id') \
            .eq('season_id', season_id).eq('tegenstander', opp).eq('datum', datum) \
            .maybe_single().execute()
        if res.data:
            col_to_match[c] = res.data['id']

    count = 0
    for row in range(4, ws.max_row + 1):
        if not isinstance(ws.cell(row, 1).value, (int, float)):
            continue
        naam_raw = ws.cell(row, 2).value
        if not naam_raw:
            continue
        naam = str(naam_raw).strip()
        if not naam:
            continue
        player_id = get_or_create_player(naam)
        for c, match_id in col_to_match.items():
            val = ws.cell(row, c).value
            if val is None:
                continue
            try:
                aantal = int(float(val))
            except (ValueError, TypeError):
                continue
            if aantal <= 0:
                continue
            sb.table('assists').insert({
                'match_id': match_id, 'player_id': player_id, 'aantal': aantal,
            }).execute()
            count += 1

    print(f"    → {count} assist records.")


# ── Import cards (Kaarten) sheet ──────────────────────────────
def import_cards_sheet(ws, season_name: str):
    season_id = get_or_create_season(season_name)
    print(f"  Importing cards for {season_name}…")

    col_to_match = {}
    for c in range(3, ws.max_column + 1):
        opp_val = ws.cell(2, c).value
        date_raw = ws.cell(3, c).value
        if not opp_val:
            continue
        opp = str(opp_val).strip()
        datum = parse_date(date_raw, season_name) or '1900-01-01'
        res = sb.table('matches').select('id') \
            .eq('season_id', season_id).eq('tegenstander', opp).eq('datum', datum) \
            .maybe_single().execute()
        if res.data:
            col_to_match[c] = res.data['id']

    count = 0
    for row in range(4, ws.max_row + 1):
        if not isinstance(ws.cell(row, 1).value, (int, float)):
            continue
        naam_raw = ws.cell(row, 2).value
        if not naam_raw:
            continue
        naam = str(naam_raw).strip()
        if not naam:
            continue
        player_id = get_or_create_player(naam)
        for c, match_id in col_to_match.items():
            val = ws.cell(row, c).value
            if val is None:
                continue
            try:
                aantal = int(float(val))
            except (ValueError, TypeError):
                continue
            if aantal <= 0:
                continue
            sb.table('kaarten').insert({
                'match_id': match_id, 'player_id': player_id, 'geel': aantal,
            }).execute()
            count += 1

    print(f"    → {count} card records.")


# ── Import Men of the Match sheet ─────────────────────────────
def import_motm_sheet(ws, season_name: str):
    season_id = get_or_create_season(season_name)
    print(f"  Importing MOTM for {season_name}…")

    col_to_match = {}
    for c in range(3, ws.max_column + 1):
        opp_val = ws.cell(2, c).value
        date_raw = ws.cell(3, c).value
        if not opp_val:
            continue
        opp = str(opp_val).strip()
        datum = parse_date(date_raw, season_name) or '1900-01-01'
        res = sb.table('matches').select('id') \
            .eq('season_id', season_id).eq('tegenstander', opp).eq('datum', datum) \
            .maybe_single().execute()
        if res.data:
            col_to_match[c] = res.data['id']

    count = 0
    for row in range(4, ws.max_row + 1):
        if not isinstance(ws.cell(row, 1).value, (int, float)):
            continue
        naam_raw = ws.cell(row, 2).value
        if not naam_raw:
            continue
        naam = str(naam_raw).strip()
        if not naam:
            continue
        player_id = get_or_create_player(naam)
        for c, match_id in col_to_match.items():
            val = ws.cell(row, c).value
            if val is None:
                continue
            try:
                if int(float(val)) >= 1:
                    # upsert so only one MOTM per match
                    sb.table('motm').upsert(
                        {'match_id': match_id, 'player_id': player_id},
                        on_conflict='match_id'
                    ).execute()
                    count += 1
            except (ValueError, TypeError):
                continue

    print(f"    → {count} MOTM records.")


# ── Main ──────────────────────────────────────────────────────
def main():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Sheets found: {wb.sheetnames}\n")

    # ── Step 1: Wipe all existing match-related data ──────────
    # (matches CASCADE → aanwezigheid, doelpunten, assists, kaarten, motm)
    print("⚠  Deleting all existing match data (CASCADE)…")
    # Delete in dependency order
    sb.table('motm').delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
    sb.table('kaarten').delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
    sb.table('assists').delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
    sb.table('doelpunten').delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
    sb.table('aanwezigheid').delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
    sb.table('matches').delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
    print("  ✓ Existing match data cleared.\n")

    # ── Step 2: Import each season attendance sheet ───────────
    for season_name in SEASON_SHEETS:
        if season_name not in wb.sheetnames:
            print(f"  ⚠ Sheet '{season_name}' not found, skipping.")
            continue
        print(f"\n=== Seizoen {season_name} ===")
        import_season_sheet(wb[season_name], season_name)

    # ── Step 3: Import 2025-2026 goals / assists / cards / MOTM
    print(f"\n=== Goals / Assists / Cards / MOTM for {CURRENT_SEASON} ===")

    goals_sheet = f'Doelpunten {CURRENT_SEASON}'
    if goals_sheet in wb.sheetnames:
        import_goals_sheet(wb[goals_sheet], CURRENT_SEASON)
    else:
        print(f"  ⚠ Sheet '{goals_sheet}' not found.")

    assists_sheet = f'Assists {CURRENT_SEASON}'
    if assists_sheet in wb.sheetnames:
        import_assists_sheet(wb[assists_sheet], CURRENT_SEASON)
    else:
        print(f"  ⚠ Sheet '{assists_sheet}' not found.")

    cards_sheet = f'Kaarten {CURRENT_SEASON}'
    if cards_sheet in wb.sheetnames:
        import_cards_sheet(wb[cards_sheet], CURRENT_SEASON)
    else:
        print(f"  ⚠ Sheet '{cards_sheet}' not found.")

    motm_sheet = f'Men of the Match {CURRENT_SEASON}'
    if motm_sheet in wb.sheetnames:
        import_motm_sheet(wb[motm_sheet], CURRENT_SEASON)
    else:
        print(f"  ⚠ Sheet '{motm_sheet}' not found.")

    print("\n✓ Import completed!")


if __name__ == '__main__':
    main()
